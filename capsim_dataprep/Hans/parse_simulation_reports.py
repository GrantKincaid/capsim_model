
from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path
from typing import Dict, List, Tuple, Any

from openpyxl import load_workbook

CATEGORY_SHEETS = {
    "Traditional": 0,
    "Low End": 1,
    "High End": 2,
    "Performance": 3,
    "Size": 4,
}

PRODUCT_COLUMNS = {
    "Name": "product_name",
    "Price": "price",
    "Units Sold": "units_sold",
    "Potential Sold": "potential_sold",
    "Stock Out": "stock_out",
    "Age": "age",
    "Performance": "performance",
    "Size": "size",
    "Reliability": "reliability",
    "Sales Budget": "sales_budget",
    "Customer Accessibility": "customer_accessibility",
    "Promo Budget": "promo_budget",
    "Customer Awareness": "customer_awareness",
    "Customer Satisfaction": "customer_satisfaction",
}


def safe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "").replace("$", "")
    if s.endswith("%"):
        try:
            return float(s[:-1]) / 100.0
        except ValueError:
            return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_midpoint_range(text: str) -> float | None:
    nums = re.findall(r"[-+]?\d[\d,]*\.?\d*", str(text))
    vals = [safe_float(x) for x in nums]
    vals = [x for x in vals if x is not None]
    if not vals:
        return None
    if len(vals) == 1:
        return vals[0]
    return sum(vals[:2]) / 2.0


def parse_positioning(text: str) -> Tuple[float | None, float | None]:
    perf_match = re.search(r"Performance\s+([-+]?\d*\.?\d+)", str(text), flags=re.I)
    size_match = re.search(r"Size\s+([-+]?\d*\.?\d+)", str(text), flags=re.I)
    perf = float(perf_match.group(1)) if perf_match else None
    size = float(size_match.group(1)) if size_match else None
    return perf, size


def find_product_header_row(ws) -> int:
    for r in range(1, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, 15)]
        if values[:4] == ["Name", "Price", "Units Sold", "Potential Sold"]:
            return r
    raise ValueError(f"Could not find product table header in sheet {ws.title!r}")


def parse_sheet_targets(ws) -> Dict[str, float | None]:
    targets = {
        "age_target": None,
        "price_target_mid": None,
        "performance_target": None,
        "size_target": None,
        "reliability_target_mid": None,
        "segment_total_market_size": None,
        "segment_total_units_sold": None,
        "segment_pct_total_industry": None,
        "next_year_growth_rate": None,
    }

    for r in range(1, min(ws.max_row, 20) + 1):
        label = ws.cell(r, 1).value
        value = ws.cell(r, 2).value
        if label == "Age":
            parsed_age_target = parse_midpoint_range(value)
            if parsed_age_target is not None and parsed_age_target < 0.5:
                parsed_age_target = 0.5
            targets["age_target"] = parsed_age_target
        elif label == "Price":
            targets["price_target_mid"] = parse_midpoint_range(value)
        elif label == "Positioning":
            perf_target, size_target = parse_positioning(value)
            targets["performance_target"] = perf_target
            targets["size_target"] = size_target
        elif label == "Reliability":
            targets["reliability_target_mid"] = parse_midpoint_range(value)
        elif isinstance(label, str) and "Total Market Size" in label:
            targets["segment_total_market_size"] = safe_float(value)
        elif isinstance(label, str) and "Total Units Sold" in label:
            targets["segment_total_units_sold"] = safe_float(value)
        elif label == "Segment % of Total Industry":
            targets["segment_pct_total_industry"] = safe_float(value)
        elif isinstance(label, str) and "Demand Growth Rate" in label:
            targets["next_year_growth_rate"] = safe_float(value)

    return targets


def parse_products(ws) -> Dict[str, Dict[str, Any]]:
    header_row = find_product_header_row(ws)
    headers = [ws.cell(header_row, c).value for c in range(1, 15)]
    col_map = {idx + 1: PRODUCT_COLUMNS.get(h, h) for idx, h in enumerate(headers)}

    products: Dict[str, Dict[str, Any]] = {}
    r = header_row + 1
    while r <= ws.max_row:
        name = ws.cell(r, 1).value
        if name in (None, ""):
            r += 1
            continue
        row_data: Dict[str, Any] = {}
        for c in range(1, 15):
            key = col_map.get(c)
            row_data[key] = ws.cell(r, c).value
        products[str(name).strip()] = row_data
        r += 1
    return products


def parse_round_year_from_filename(path: Path) -> Tuple[int | None, int | None]:
    round_match = re.search(r"round[_ -]?(\d+)", path.stem, flags=re.I)
    year_match = re.search(r"(20\d{2})", path.stem)
    round_num = int(round_match.group(1)) if round_match else None
    year = int(year_match.group(1)) if year_match else None
    return round_num, year


def build_master_product_map(sheet_products: Dict[str, Dict[str, Dict[str, Any]]]) -> Dict[str, Dict[str, Any]]:
    master: Dict[str, Dict[str, Any]] = {}
    for sheet_name in CATEGORY_SHEETS:
        for product_name, row in sheet_products[sheet_name].items():
            if product_name not in master:
                master[product_name] = row.copy()
                master[product_name]["source_sheet_for_defaults"] = sheet_name
    return master


def pct_of_target(value: float | None, target: float | None) -> float | None:
    if value is None or target in (None, 0):
        return None
    return (value / target) - 1 


def workbook_to_rows(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    sheet_targets: Dict[str, Dict[str, Any]] = {}
    sheet_products: Dict[str, Dict[str, Dict[str, Any]]] = {}

    for sheet_name in CATEGORY_SHEETS:
        ws = wb[sheet_name]
        sheet_targets[sheet_name] = parse_sheet_targets(ws)
        sheet_products[sheet_name] = parse_products(ws)

    all_products = sorted({p for sheet_map in sheet_products.values() for p in sheet_map})
    master_products = build_master_product_map(sheet_products)
    round_num, year = parse_round_year_from_filename(path)

    rows: List[Dict[str, Any]] = []

    for category_name, category_code in CATEGORY_SHEETS.items():
        cat_targets = sheet_targets[category_name]
        segment_total_market_size = cat_targets["segment_total_market_size"] or 0.0
        for product_name in all_products:
            in_category = product_name in sheet_products[category_name]
            base_row = sheet_products[category_name].get(product_name, master_products[product_name])

            price = safe_float(base_row.get("price"))
            age = safe_float(base_row.get("age"))
            performance = safe_float(base_row.get("performance"))
            size = safe_float(base_row.get("size"))
            reliability = safe_float(base_row.get("reliability"))

            # For products not explicitly present in this category table:
            # - carry forward product attributes from the product's available row
            # - set demand/result fields to zero
            units_sold = safe_float(base_row.get("units_sold")) if in_category else 0.0
            potential_sold = safe_float(base_row.get("potential_sold")) if in_category else 0.0
            stock_out = base_row.get("stock_out") if in_category else "No Data"
            customer_accessibility = safe_float(base_row.get("customer_accessibility"))
            customer_awareness = safe_float(base_row.get("customer_awareness"))
            sales_budget = safe_float(base_row.get("sales_budget"))
            promo_budget = safe_float(base_row.get("promo_budget"))
            customer_satisfaction = safe_float(base_row.get("customer_satisfaction"))

            row = {
                "source_file": path.name,
                "round": round_num,
                "year": year,
                "category_name": category_name,
                "category_code": category_code,
                "product_name": product_name,
                "product_seen_in_category_tab": int(in_category),
                "default_source_sheet": base_row.get("source_sheet_for_defaults", category_name),
                "segment_total_market_size": segment_total_market_size,
                "segment_total_units_sold": cat_targets["segment_total_units_sold"],
                "segment_pct_total_industry": cat_targets["segment_pct_total_industry"],
                "next_year_growth_rate": cat_targets["next_year_growth_rate"],
                "units_sold": units_sold,
                "potential_sold": potential_sold,
                "units_sold_share_of_segment": (units_sold / segment_total_market_size) if segment_total_market_size else 0.0,
                "target_potential_share_of_segment": (potential_sold / segment_total_market_size) if segment_total_market_size else 0.0,
                "stock_out": stock_out,
                "price": price,
                "age": age,
                "performance": performance,
                "size": size,
                "reliability": reliability,
                "sales_budget": sales_budget,
                "promo_budget": promo_budget,
                "customer_accessibility": customer_accessibility,
                "customer_awareness": customer_awareness,
                "customer_satisfaction": customer_satisfaction,
                "age_pct_of_target": pct_of_target(age, cat_targets["age_target"]),
                "price_pct_of_target_mid": pct_of_target(price, cat_targets["price_target_mid"]),
                "performance_pct_of_target": pct_of_target(performance, cat_targets["performance_target"]),
                "size_pct_of_target": pct_of_target(size, cat_targets["size_target"]),
                "reliability_pct_of_target_mid": pct_of_target(reliability, cat_targets["reliability_target_mid"]),
            }
            rows.append(row)

    return rows


def write_csv(rows: List[Dict[str, Any]], output_path: Path) -> None:
    if not rows:
        raise ValueError("No rows generated.")
    fieldnames = list(rows[0].keys())
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert Capsim-style Excel simulation reports into one ML-ready CSV.")
    parser.add_argument("input_folder", help="Folder containing .xlsx simulation reports")
    parser.add_argument("-o", "--output", default="simulation_ml_dataset.csv", help="Output CSV path")
    parser.add_argument("--glob", default="*.xlsx", help="Glob pattern for input files")
    args = parser.parse_args()

    input_folder = Path(args.input_folder)
    files = sorted(input_folder.glob(args.glob))
    if not files:
        raise FileNotFoundError(f"No files matching {args.glob!r} found in {input_folder}")

    all_rows: List[Dict[str, Any]] = []
    for path in files:
        all_rows.extend(workbook_to_rows(path))

    write_csv(all_rows, Path(args.output))
    print(f"Wrote {len(all_rows)} rows to {args.output}")


if __name__ == "__main__":
    main()
